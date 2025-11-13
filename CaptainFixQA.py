from openpyxl import load_workbook
import json
import subprocess
import sys
from langchain_core.prompts import ChatPromptTemplate
import pandas as pd
from selenium import webdriver
from langchain_openai import ChatOpenAI



def analyze_html_with_llm(html_content,llm):

    template ='''אני נותן לך /וד HTML של דף אינטרנט
    {html}
    זהה את האלמנטים הפעילים (כפתורים, קישורים, טפסים, שדות קלט) ותאר את הפעולות האפשריות שניתן לבצע עליהם
    החזר את זה כרשימה ברורה, ממוינת לפי סוג האלמנט
     '''

    prompt = ChatPromptTemplate.from_template(template)
    chain = prompt | llm
    response = chain.invoke({"html":html_content})

    return response.content


#generate a test cases
def generate_test_cases_by_llm(html_content:str , elements:str , llm):
    #depth = number of test cases
    depth = CONFIG.get("depth", "10-12 test cases in total")
    template =  """
        HTML: {html}
        ELEMENTS: {elements}   # e.g. ["/", "/login", "/contact", "/products"]
        Constraints:
        - Generate Suites: (Smoke, Navigation, Forms).
        - Produce: {depth}  (number of test cases total)
        - Each test case must include: id, title, priority (P0,P1,P2), steps (action/selector/value), expected(very important), tags (optional).
        - Allowed actions: goto, click, fill, assert_text, assert_element, assert_url, assert_status, wait, select, upload, hover, back, forward.
        - Use selectors that are likely to exist (ids, data-test attributes). If no selector possible, use the page-level step action (e.g., "assert_text" with selector "body").
        - Every step must be a small atomic action (one action per step).
        - Use consistent case for ids: <SUITE>-C<number> (e.g., SMOKE-C1).
        - Output must validate as JSON. Do not wrap the JSON in markdown or code fences.
        - If a page from the provided list doesn't exist, include a test case to assert 200 status for that page.
        - Please generate the full JSON test plan with all suites and cases in one reply, don’t truncate.
        Return the TestPlan JSON now.
    """
    prompt = ChatPromptTemplate.from_template(template)
    chain = prompt | llm


    try:
        response = chain.invoke({"html":html_content , "elements":elements , "depth":depth})
    except Exception as e:
        raise RuntimeError(f"Failed to invoke chain: {e}")
    if hasattr(response, "content"):
        test_cases = response.content
    else:
        test_cases = str(response)

    test_cases = json.loads(test_cases)

    return test_cases

# export json file
def export_testplan_json_file(filename: str , content: dict):
    with open(filename, "w" , encoding="utf-8") as f:
        json.dump(content , f , indent=4 , ensure_ascii=False)

def export_testplan_to_excel( filename: str , content: dict):
    """
    Export testplan JSON -> Excel workbook with two sheets:
      - TestCases: one row per test case (high-level)
      - Steps: one row per step (detailed)
    """
    # Flatten test cases
    case_rows = []
    step_rows = []
    for suite in content.get("suites", []):
        suite_name = suite.get("name", "")
        for case in suite.get("cases", []):
            case_id = case.get("id", "")
            title = case.get("title", "")
            priority = case.get("priority", "")
            expected = case.get("expected", "")
            tags = ", ".join(case.get("tags", []))
            # Build steps text (multi-line)
            steps = case.get("steps", [])
            step_texts = []
            for i, s in enumerate(steps, start=1):
                # Represent step succinctly
                action = s.get("action", "")
                selector = s.get("selector") or s.get("value") or ""
                value = s.get("value") if "value" in s and s.get("selector") else ""
                # prefer explicit fields
                if s.get("selector") and "value" in s:
                    step_text = f"{i}. {action} selector={s.get('selector')} value={s.get('value')}"
                elif s.get("selector"):
                    step_text = f"{i}. {action} selector={s.get('selector')}"
                elif "value" in s:
                    step_text = f"{i}. {action} value={s.get('value')}"
                else:
                    step_text = f"{i}. {action}"
                step_texts.append(step_text)

                # add to detailed steps sheet
                step_rows.append({
                    "Suite": suite_name,
                    "Case ID": case_id,
                    "Step No": i,
                    "Action": action,
                    "Selector": s.get("selector", ""),
                    "Value": s.get("value", ""),
                    "Raw": json.dumps(s, ensure_ascii=False)
                })

            steps_cell = "\n".join(step_texts)
            case_rows.append({
                "Suite": suite_name,
                "Case ID": case_id,
                "Title": title,
                "Priority": priority,
                "Expected": expected,
                "Tags": tags,
                "Steps": steps_cell
            })

    # Create DataFrames
    df_cases = pd.DataFrame(case_rows)
    df_steps = pd.DataFrame(step_rows)

    # Write to Excel with two sheets
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df_cases.to_excel(writer, sheet_name="TestCases", index=False)
        df_steps.to_excel(writer, sheet_name="Steps", index=False)


    # Auto-adjust column widths (openpyxl)
    try:

        wb = load_workbook(filename)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # determine max width per column
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(cell_value))
            for col, value in dims.items():
                adjusted = min(value + 2, 60)  # cap width
                ws.column_dimensions[col].width = adjusted
        wb.save(filename)
    except Exception as e:
        # Non-fatal: still valid Excel, just not auto-sized
        print("Warning: could not auto-size columns:", e)


def exec_selenium_script(selenium_script: str, filename: str = "seleniumtest.py"):
    """
    Save selenium_script into filename and execute it with the current Python interpreter.
    """
    try:
        # Write script to file
        with open(filename, "w", encoding="utf-8") as f:
            f.write(selenium_script)

        print(f"[INFO] Script written to {filename}")

        # Run the file using the same Python interpreter
        result = subprocess.run(
            [sys.executable, filename],
            capture_output=True,
            text=True
        )

        # Print stdout & stderr
        if result.stdout:
            print("[STDOUT]:")
            print(result.stdout)

        if result.stderr:
            print("[STDERR]:")
            print(result.stderr)

        return result.returncode

    except Exception as e:
        print(f"[ERROR] Failed to execute {filename}: {e}")
        return -1

def main():
    if len(sys.argv) > 1:
        url = str(sys.argv[1])
    else:
        url = "file:///Users/raniaburaia/Desktop/certificate/Captain-Fix/User_Managment.html"

    driver,llm = set_up()

    # Access config values anywhere:
    # test_plan_report = CONFIG.get("testPlanReport","JSON")

    driver.get(url)


    html_content = driver.page_source
    elements = analyze_html_with_llm(html_content, llm)
    print(elements)

    driver.quit()

    #----driver.get("https://www.w3schools.com/html/html_forms.asp")
    #----fname = driver.find_element(By.ID,"fname")

    # test_cases = generate_test_cases_by_llm(html_content,elements , llm)
    #
    # # If user requested JSON report, write it
    # if test_plan_report and ("JSON" in test_plan_report or "JSON" in (test_plan_report.upper())):
    #     export_testplan_json_file("PlanJson.json", test_cases)
    # if test_plan_report and ("EXCEL" in test_plan_report or "EXCEL" in test_plan_report.upper()):
    #     export_testplan_to_excel("planExcel.xlsx", test_cases)
    #
    # # generate a selenium script to check the web using test_cases file and exec the file
    # selenium_script = create_selenium_script(test_cases,html_content, llm)
    # exec_selenium_script(selenium_script)
    #




def analyze_html_with_llm(html_content,llm):

    template ='''אני נותן לך /וד HTML של דף אינטרנט
    {html}
    זהה את האלמנטים הפעילים (כפתורים, קישורים, טפסים, שדות קלט) ותאר את הפעולות האפשריות שניתן לבצע עליהם
    החזר את זה כרשימה ברורה, ממוינת לפי סוג האלמנט
     '''

    prompt = ChatPromptTemplate.from_template(template)
    chain = prompt | llm
    response = chain.invoke({"html":html_content})

    return response.content











# generate a test cases
def generate_test_cases_by_llm(html_content:str , elements:str , llm):
    #depth = number of test cases
    depth = CONFIG.get("depth", "10-12 test cases in total")
    template =  """
        HTML: {html}
        ELEMENTS: {elements}   # e.g. ["/", "/login", "/contact", "/products"]
        Constraints:
        - Generate Suites: (Smoke, Navigation, Forms).
        - Produce: {depth}  (number of test cases total)
        - Each test case must include: id, title, priority (P0,P1,P2), steps (action/selector/value), expected(very important), tags (optional).
        - Allowed actions: goto, click, fill, assert_text, assert_element, assert_url, assert_status, wait, select, upload, hover, back, forward.
        - Use selectors that are likely to exist (ids, data-test attributes). If no selector possible, use the page-level step action (e.g., "assert_text" with selector "body").
        - Every step must be a small atomic action (one action per step).
        - Use consistent case for ids: <SUITE>-C<number> (e.g., SMOKE-C1).
        - Output must validate as JSON. Do not wrap the JSON in markdown or code fences.
        - If a page from the provided list doesn't exist, include a test case to assert 200 status for that page.
        - Please generate the full JSON test plan with all suites and cases in one reply, don’t truncate.
        Return the TestPlan JSON now.
    """
    prompt = ChatPromptTemplate.from_template(template)
    chain = prompt | llm


    try:
        response = chain.invoke({"html":html_content , "elements":elements , "depth":depth})
    except Exception as e:
        raise RuntimeError(f"Failed to invoke chain: {e}")
    if hasattr(response, "content"):
        test_cases = response.content
    else:
        test_cases = str(response)

    test_cases = json.loads(test_cases)

    return test_cases

# export json file
def export_testplan_json_file(filename: str , content: dict):
    with open(filename, "w" , encoding="utf-8") as f:
        json.dump(content , f , indent=4 , ensure_ascii=False)

def export_testplan_to_excel( filename: str , content: dict):
    """
    Export testplan JSON -> Excel workbook with two sheets:
      - TestCases: one row per test case (high-level)
      - Steps: one row per step (detailed)
    """
    # Flatten test cases
    case_rows = []
    step_rows = []
    for suite in content.get("suites", []):
        suite_name = suite.get("name", "")
        for case in suite.get("cases", []):
            case_id = case.get("id", "")
            title = case.get("title", "")
            priority = case.get("priority", "")
            expected = case.get("expected", "")
            tags = ", ".join(case.get("tags", []))
            # Build steps text (multi-line)
            steps = case.get("steps", [])
            step_texts = []
            for i, s in enumerate(steps, start=1):
                # Represent step succinctly
                action = s.get("action", "")
                selector = s.get("selector") or s.get("value") or ""
                value = s.get("value") if "value" in s and s.get("selector") else ""
                # prefer explicit fields
                if s.get("selector") and "value" in s:
                    step_text = f"{i}. {action} selector={s.get('selector')} value={s.get('value')}"
                elif s.get("selector"):
                    step_text = f"{i}. {action} selector={s.get('selector')}"
                elif "value" in s:
                    step_text = f"{i}. {action} value={s.get('value')}"
                else:
                    step_text = f"{i}. {action}"
                step_texts.append(step_text)

                # add to detailed steps sheet
                step_rows.append({
                    "Suite": suite_name,
                    "Case ID": case_id,
                    "Step No": i,
                    "Action": action,
                    "Selector": s.get("selector", ""),
                    "Value": s.get("value", ""),
                    "Raw": json.dumps(s, ensure_ascii=False)
                })

            steps_cell = "\n".join(step_texts)
            case_rows.append({
                "Suite": suite_name,
                "Case ID": case_id,
                "Title": title,
                "Priority": priority,
                "Expected": expected,
                "Tags": tags,
                "Steps": steps_cell
            })

    # Create DataFrames
    df_cases = pd.DataFrame(case_rows)
    df_steps = pd.DataFrame(step_rows)

    # Write to Excel with two sheets
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df_cases.to_excel(writer, sheet_name="TestCases", index=False)
        df_steps.to_excel(writer, sheet_name="Steps", index=False)


    # Auto-adjust column widths (openpyxl)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # determine max width per column
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(cell_value))
            for col, value in dims.items():
                adjusted = min(value + 2, 60)  # cap width
                ws.column_dimensions[col].width = adjusted
        wb.save(filename)
    except Exception as e:
        # Non-fatal: still valid Excel, just not auto-sized
        print("Warning: could not auto-size columns:", e)

def create_selenium_script(test_cases, html_content, llm):
    template = """
    You are a code generator. Output EXACTLY one Python 3 script (plain text ONLY — no JSON, no commentary, no markdown fences, no extra whitespace outside the script). 
    The script must be runnable as-is and implement the test cases provided.
    Requirements for the produced Python script:
    1. Output format & content: 
    	- The LLM should output ONLY the Python source code. Nothing else. Do not include surrounding backticks or explanation.
    	- The script must be syntactically valid Python 3. Ensure ALL parentheses, brackets, braces and quotes are balanced , make sure all the ( ) is right. 
    2. Top-level behavior: 
    	- Launch Selenium Chrome WebDriver using webdriver.Chrome() (assume chromedriver is available on PATH). 
    	- Use WebDriverWait with a sensible timeout (8 seconds by default).
    	- Execute each test suite and case from the provided {test_cases} input.
    	- For each step, follow the mapping described below to actual Selenium operations. 
    	- Each case must be wrapped in try/except to capture failures; continue running other cases after failures. 
    	- When a browser confirm/alert appears, accept it automatically.
    3. Action -> Selenium mapping: 
    	- "goto": * open the local file {url}. 
    	- "fill": * Locate element with By.CSS_SELECTOR using explicit wait presence_of_element_located. Then clear() and send_keys(value). 
    	- "click": * Locate element with By.CSS_SELECTOR using explicit wait for element to be clickable, then click(). 
    	- "assert_element": * If value == "visible" wait until visibility_of_element_located. If value == "hidden" wait until invisibility_of_element_located. 
    	- "assert_text": * Wait until the element located by selector contains the expected text (partial match acceptable). 
    	- Unknown action: mark step as skipped and record a warning in results. 
    5. Implementation notes: 
    	- Use imports: json, os, traceback, time, and selenium required modules (webdriver, By, WebDriverWait, expected_conditions as EC). 
    	- Use driver.quit() in a finally block. - Be defensive about missing selectors: if an element can't be located within the timeout, mark that step failed with the appropriate error and continue to the next step/case. 

    6. Error handling / structure***
        - Implement helper functions:
          * write_html_to_tempfile(html) -> file_path
          * accept_alert_if_present(driver, timeout=2) -> bool
          * screenshot(driver, prefix) -> path
          * run_step(driver, step) -> bool
          * run_case(driver, case) -> list[dict]
          * run_suite(driver, suite) -> dict
        - Each step MUST be wrapped in try/except catching TimeoutException, UnexpectedAlertPresentException, NoSuchElementException, Exception. On exception:
          * print an error with suite/case/title/step info
          * take screenshot and include path in logs
          * return False for that step and continue

    7. Final output
        - After run, print a JSON summary of suites->cases->steps with status 'passed'/'failed'/'skipped' and any screenshot paths write it on file call it result.json , make it easy to read .
        - No extra prints outside the script other than the script's own logging.

    8. Use the provided inputs: 
    	- {html} will contain the full HTML content for the page. 
    	- {test_cases} will contain the test suites JSON. 
    9. Keep the code reasonably compact and readable (functions like run_step, run_case, run_suite are encouraged). 

    """

    # Build the LangChain prompt and call the LLM chain
    prompt = ChatPromptTemplate.from_template(template)
    chain = prompt | llm

    try:
        response = chain.invoke({"html": html_content, "test_cases": test_cases,"url":CONFIG.get("url","file:////Users/raniaburaia/Desktop/certificate/Captain-Fix/User_Managment.html")})
    except Exception as e:
        raise RuntimeError(f"Failed to invoke chain: {e}")

    if hasattr(response, "content"):
        selenium_script = response.content
    else:
        selenium_script = str(response)
    return selenium_script

def exec_selenium_script(selenium_script: str, filename: str = "seleniumtest.py"):
    """
    Save selenium_script into filename and execute it with the current Python interpreter.
    """
    try:
        # Write script to file
        with open(filename, "w", encoding="utf-8") as f:
            f.write(selenium_script)

        print(f"[INFO] Script written to {filename}")

        # Run the file using the same Python interpreter
        result = subprocess.run(
            [sys.executable, filename],
            capture_output=True,
            text=True
        )

        # Print stdout & stderr
        if result.stdout:
            print("[STDOUT]:")
            print(result.stdout)

        if result.stderr:
            print("[STDERR]:")
            print(result.stderr)

        return result.returncode

    except Exception as e:
        print(f"[ERROR] Failed to execute {filename}: {e}")
        return -1


def main():
    if len(sys.argv) > 1:
        url = sys.argv[1]
    else:
        url = "file:////Users/raniaburaia/Desktop/certificate/Captain-Fix/User_Managment.html"

    llm = ChatOpenAI(
        model='gpt-3.5-turbo',
        temperature=0.1,
        openai_api_key="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"
    )

    driver = webdriver.Chrome()
    driver.get(url)
    driver.maximize_window()
    html_content = driver.page_source
    elements = analyze_html_with_llm(html_content,llm)
    print(elements)
    driver.quit()

    #test_plan_report = CONFIG.get("testPlanReport", "JSON")
    #test_cases = generate_test_cases_by_llm(html_content, elements, llm)
    # # If user requested JSON report, write it
    # if test_plan_report and ("JSON" in test_plan_report or "JSON" in (test_plan_report.upper())):
    #     export_testplan_json_file("PlanJson.json", test_cases)
    # if test_plan_report and ("EXCEL" in test_plan_report or "EXCEL" in test_plan_report.upper()):
    #     export_testplan_to_excel("planExcel.xlsx", test_cases)
    #
    # # generate a selenium script to check the web using test_cases file and exec the file
    # selenium_script = create_selenium_script(test_cases,html_content, llm)
    # exec_selenium_script(selenium_script)


if __name__ == "__main__":
    main()

